import os
import io
import asyncio
import logging
import json
import re
import pandas as pd
import yfinance as yf
import numpy as np
import plotly.graph_objects as go
import chainlit as cl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from llama_index.core import VectorStoreIndex, SimpleDirectoryReader, Settings
from llama_index.core.agent.workflow import AgentWorkflow, ReActAgent
from llama_index.core.tools import QueryEngineTool, ToolMetadata, FunctionTool
from llama_index.core.schema import Document
from llama_index.llms.anthropic import Anthropic
from llama_index.embeddings.huggingface import HuggingFaceEmbedding
from llama_index.tools.tavily_research import TavilyToolSpec

# ---------------------------------------------------------------------------
# 1. LOGGING & STARTUP VALIDATION
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("PortfolioAI")


def _validate_env():
    missing = [
        var for var in ["ANTHROPIC_API_KEY", "TAVILY_API_KEY"]
        if not os.environ.get(var)
    ]
    if missing:
        raise EnvironmentError(
            f"Missing required environment variables: {', '.join(missing)}. "
            "Ensure they are set in your .env file or Railway dashboard."
        )


# ---------------------------------------------------------------------------
# 2. TICKER RESOLUTION  (company name → ticker symbol)
# ---------------------------------------------------------------------------
def resolve_ticker(input_str: str) -> tuple[str, str]:
    """
    Accept a ticker or company name and return (ticker, company_name).
    Uses yfinance search; falls back to treating the input as a raw ticker.
    """
    input_str = input_str.strip().upper()
    try:
        search = yf.Search(input_str, max_results=1)
        quotes = search.quotes
        if quotes:
            ticker = quotes[0].get("symbol", input_str)
            name = quotes[0].get("longname") or quotes[0].get("shortname") or ticker
            return ticker, name
    except Exception:
        pass
    # Fallback: treat as raw ticker and look up name
    try:
        info = yf.Ticker(input_str).info
        name = info.get("longName") or info.get("shortName") or input_str
        return input_str, name
    except Exception:
        return input_str, input_str


def parse_ticker_input(raw: str) -> list[str]:
    """
    Parse a free-text string into a list of ticker/company tokens.
    Handles comma, space, semicolon, and newline separators.
    """
    tokens = re.split(r"[,;\n]+", raw)
    return [t.strip() for t in tokens if t.strip()]


def parse_file_upload(file_bytes: bytes, filename: str) -> list[str]:
    """
    Read a CSV or Excel file and extract all non-empty values from
    the first column (or a column named 'ticker'/'symbol'/'company').
    """
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            df = pd.read_excel(io.BytesIO(file_bytes))

        # Try to find a named column first
        col = None
        for c in df.columns:
            if c.strip().lower() in ("ticker", "symbol", "company", "name", "equity"):
                col = c
                break
        if col is None:
            col = df.columns[0]

        values = df[col].dropna().astype(str).str.strip().tolist()
        return [v for v in values if v]
    except Exception as exc:
        logger.error("File parse error: %s", exc)
        return []


# ---------------------------------------------------------------------------
# 3. QUANT METRICS  (direct — not via agent, for structured data capture)
# ---------------------------------------------------------------------------
def get_quant_metrics(ticker: str, risk_free_rate: float = 0.04) -> dict:
    """Return structured quant metrics for a ticker."""
    try:
        stock = yf.Ticker(ticker)
        hist = stock.history(period="1y")
        info = stock.info

        if hist.empty:
            return {"error": f"No price data for {ticker}", "signal": "negative"}

        returns = hist["Close"].pct_change().dropna()
        ann_return = (hist["Close"].iloc[-1] / hist["Close"].iloc[0]) - 1
        ann_vol = returns.std() * np.sqrt(252)
        sharpe = (ann_return - risk_free_rate) / ann_vol if ann_vol != 0 else 0

        return {
            "ticker": ticker,
            "current_price": f"${info.get('currentPrice', hist['Close'].iloc[-1]):.2f}",
            "market_cap": _fmt_large(info.get("marketCap")),
            "pe_ratio": info.get("trailingPE", "N/A"),
            "52w_high": f"${info.get('fiftyTwoWeekHigh', 'N/A')}",
            "52w_low": f"${info.get('fiftyTwoWeekLow', 'N/A')}",
            "ann_return": f"{ann_return:.2%}",
            "ann_volatility": f"{ann_vol:.2%}",
            "sharpe_ratio": round(sharpe, 2),
            "signal": "positive" if sharpe > 1.0 else "negative",
        }
    except Exception as exc:
        logger.error("Quant metrics failed for %s: %s", ticker, exc)
        return {"error": str(exc), "signal": "negative"}


def _fmt_large(n) -> str:
    if n is None:
        return "N/A"
    if n >= 1e12:
        return f"${n/1e12:.2f}T"
    if n >= 1e9:
        return f"${n/1e9:.2f}B"
    if n >= 1e6:
        return f"${n/1e6:.2f}M"
    return f"${n:,.0f}"


# ---------------------------------------------------------------------------
# 4. AGENTS
# ---------------------------------------------------------------------------
def get_analyst_agent(client_files_path: str = "./client_input") -> ReActAgent:
    os.makedirs(client_files_path, exist_ok=True)

    # Check for files BEFORE calling SimpleDirectoryReader
    # It throws ValueError internally if the folder is empty — guard against it
    has_files = any(
        f.is_file()
        for f in os.scandir(client_files_path)
    )
    if has_files:
        client_docs = SimpleDirectoryReader(client_files_path).load_data()
    else:
        logger.warning("No files in '%s' — analyst agent using placeholder.", client_files_path)
        client_docs = [Document(text="No private client files loaded.")]

    client_index = VectorStoreIndex.from_documents(client_docs)
    client_tool = QueryEngineTool(
        query_engine=client_index.as_query_engine(),
        metadata=ToolMetadata(
            name="client_file_search",
            description="Searches private client holdings and uploaded files.",
        ),
    )

    def search_web_10k(ticker: str, form_type: str = "10-K") -> str:
        """Fetch fundamental risks from SEC EDGAR with proper headers."""
        sec_url = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={ticker}&type={form_type}&dateb=&owner=include&count=5"
        try:
            # SEC requires a descriptive User-Agent — anonymous requests get 403
            import requests
            headers = {
                "User-Agent": "PortfolioAI research@portfolioai.com",
                "Accept-Encoding": "gzip, deflate",
                "Host": "efts.sec.gov"
            }

            # Search EDGAR full-text search for the ticker's latest 10-K
            search_url = f"https://efts.sec.gov/LATEST/search-index?q=%22{ticker}%22&dateRange=custom&startdt=2023-01-01&forms={form_type}"
            resp = requests.get(search_url, headers=headers, timeout=15)
            resp.raise_for_status()
            hits = resp.json().get("hits", {}).get("hits", [])

            if not hits:
                # Fallback: use yfinance to get business description as proxy
                info = yf.Ticker(ticker).info
                summary = (
                    f"Business: {info.get('longBusinessSummary', 'No description available.')[:500]} "
                    f"Sector: {info.get('sector', 'N/A')}. "
                    f"Industry: {info.get('industry', 'N/A')}."
                )
                return json.dumps({"summary": summary, "source_url": sec_url})

            # Pull the filing document URL from the first hit
            filing_url = hits[0].get("_source", {}).get("file_date", "")
            accession = hits[0].get("_source", {}).get("period_of_report", "")
            summary = (
                f"SEC {form_type} filing found for {ticker}. "
                f"Key risk factors should be reviewed directly at the SEC filing. "
                f"Filing period: {accession}."
            )
            return json.dumps({"summary": summary, "source_url": sec_url})

        except Exception as exc:
            logger.error("SEC filing fetch failed for %s: %s", ticker, exc)
            # Graceful fallback — use yfinance company info as substitute
            try:
                info = yf.Ticker(ticker).info
                summary = (
                    f"SEC filing unavailable. Company overview: "
                    f"{info.get('longBusinessSummary', 'No description available.')[:500]} "
                    f"Sector: {info.get('sector', 'N/A')}. "
                    f"Industry: {info.get('industry', 'N/A')}."
                )
                return json.dumps({"summary": summary, "source_url": sec_url})
            except Exception:
                return json.dumps({"summary": f"Could not retrieve filing data for {ticker}.", "source_url": sec_url})

    return ReActAgent(
        name="analyst_agent",
        description="Fundamental analyst: queries SEC EDGAR for 10-K filings.",
        system_prompt=(
            "You are a Fundamental Analyst. Use search_web_10k to get risks from SEC filings. "
            "Return a JSON object with keys: 'fundamental_summary' (string), "
            "'fundamental_signal' ('positive' or 'negative'), 'sec_url' (string)."
        ),
        tools=[client_tool, FunctionTool.from_defaults(fn=search_web_10k)],
        llm=Settings.llm,
    )


def get_pulse_agent() -> ReActAgent:
    tavily_key = os.environ.get("TAVILY_API_KEY")
    if not tavily_key:
        raise EnvironmentError("TAVILY_API_KEY is not set.")
    tavily_tool = TavilyToolSpec(api_key=tavily_key)
    return ReActAgent(
        name="pulse_agent",
        description="Pulls real-time news and sentiment via Tavily.",
        system_prompt=(
            "You are a Market Strategist. Search for recent news, analyst ratings, and "
            "insider trades for the given ticker. "
            "Return a JSON object with keys: 'news_summary' (string), "
            "'news_signal' ('positive' or 'negative'), 'news_urls' (list of up to 3 source URLs)."
        ),
        tools=tavily_tool.to_tool_list(),
        llm=Settings.llm,
    )


def fetch_news_direct(ticker: str, company_name: str) -> dict:
    """
    Fetch news directly via Tavily API without going through the agent pipeline.
    Returns structured news data reliably.
    """
    try:
        import requests
        tavily_key = os.environ.get("TAVILY_API_KEY")
        if not tavily_key:
            return {"summary": "Tavily API key not set.", "signal": "negative", "urls": []}

        resp = requests.post(
            "https://api.tavily.com/search",
            json={
                "api_key": tavily_key,
                "query": f"{company_name} {ticker} stock news analyst rating 2025",
                "search_depth": "basic",
                "max_results": 5,
                "include_answer": True,
            },
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()

        # Extract answer and URLs
        answer = data.get("answer", "") or ""
        results = data.get("results", [])
        urls = [r.get("url", "") for r in results if r.get("url")][:3]
        snippets = " ".join(r.get("content", "")[:200] for r in results[:3])

        # Simple sentiment from answer + snippets
        combined = (answer + " " + snippets).lower()
        positive_words = ["beat", "strong", "growth", "upgrade", "buy", "bullish",
                         "record", "surge", "profit", "outperform", "raised"]
        negative_words = ["miss", "weak", "decline", "downgrade", "sell", "bearish",
                         "loss", "drop", "risk", "underperform", "cut", "fell"]
        pos_score = sum(1 for w in positive_words if w in combined)
        neg_score = sum(1 for w in negative_words if w in combined)
        signal = "positive" if pos_score >= neg_score else "negative"

        summary = answer if answer else snippets[:400] if snippets else f"No recent news found for {ticker}."

        return {"summary": summary, "signal": signal, "urls": urls}

    except Exception as exc:
        logger.error("Direct Tavily fetch failed for %s: %s", ticker, exc)
        return {
            "summary": f"News fetch failed: {exc}",
            "signal": "negative",
            "urls": [],
        }


def get_quant_agent() -> ReActAgent:
    def get_metrics(ticker: str) -> str:
        return json.dumps(get_quant_metrics(ticker))

    return ReActAgent(
        name="quant_agent",
        description="Calculates Sharpe ratio and historical performance metrics.",
        tools=[FunctionTool.from_defaults(fn=get_metrics)],
        llm=Settings.llm,
    )


# ---------------------------------------------------------------------------
# 5. PORTFOLIO BOT
# ---------------------------------------------------------------------------
class PortfolioBot:

    def __init__(self):
        _validate_env()
        Settings.llm = Anthropic(
            model="claude-sonnet-4-6",
            api_key=os.environ.get("ANTHROPIC_API_KEY"),
        )
        # HuggingFace embeddings run locally — no API key needed
        Settings.embed_model = HuggingFaceEmbedding(
            model_name="BAAI/bge-small-en-v1.5"
        )
        self.output_dir = "./outputs"
        os.makedirs(self.output_dir, exist_ok=True)
        self._build_workflow()

    def _build_workflow(self):
        optimizer = ReActAgent(
            name="optimizer_agent",
            description="Lead orchestrator: coordinates all agents and synthesises results.",
            system_prompt=(
                "You are a Quant Lead analysing a single equity. You MUST:\n"
                "1. Call quant_agent with the ticker to get performance metrics and signal.\n"
                "2. Call analyst_agent with the ticker to get SEC filing risks and signal.\n"
                "3. Call pulse_agent with the ticker to get news sentiment and signal.\n"
                "4. Call calculate_consensus with the three signals.\n"
                "5. Return ONLY a valid JSON object (no markdown, no extra text) with these exact keys:\n"
                "   ticker, company_name, recommendation (BUY/HOLD/SELL), confidence (66% or 100%),\n"
                "   quant_signal, fundamental_signal, news_signal,\n"
                "   ann_return, ann_volatility, sharpe_ratio, current_price, market_cap, pe_ratio,\n"
                "   52w_high, 52w_low,\n"
                "   quant_reasoning (2-3 sentences), fundamental_reasoning (2-3 sentences),\n"
                "   news_reasoning (2-3 sentences), overall_reasoning (3-4 sentences),\n"
                "   sec_url, news_urls (list of strings), analysis_date"
            ),
            tools=[
                FunctionTool.from_defaults(fn=self.calculate_consensus),
            ],
            llm=Settings.llm,
        )
        self.workflow = AgentWorkflow(
            agents=[
                get_analyst_agent(),
                get_pulse_agent(),
                get_quant_agent(),
                optimizer,
            ],
            root_agent="optimizer_agent",
        )

    def calculate_consensus(
        self, quant_sig: str, analyst_sig: str, pulse_sig: str
    ) -> dict:
        signals = [s.strip().lower() for s in [quant_sig, analyst_sig, pulse_sig]]
        pos = signals.count("positive")
        neg = signals.count("negative")
        vote = "BUY" if pos >= 2 else "SELL"
        confidence = "100%" if (pos == 3 or neg == 3) else "66%"
        return {"recommendation": vote, "confidence": confidence}

    def _extract_json(self, text: str) -> dict | None:
        """
        Aggressively extract a JSON object from LLM output that may contain
        surrounding prose, markdown fences, or multiple JSON fragments.
        """
        # 1. Strip markdown fences
        text = re.sub(r"```(?:json)?", "", text).strip()
        text = re.sub(r"```", "", text).strip()

        # 2. Try parsing the whole thing first
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass

        # 3. Find the outermost { ... } block
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(text[start:end + 1])
            except json.JSONDecodeError:
                pass

        # 4. Try each JSON-like substring
        for match in re.finditer(r"\{[^{}]*\}", text, re.DOTALL):
            try:
                return json.loads(match.group())
            except json.JSONDecodeError:
                continue

        return None

    async def analyse_ticker(self, ticker: str, company_name: str) -> dict:
        """
        Run full analysis by calling each data source directly,
        then use LLM only to synthesise the final recommendation.
        This is more reliable than multi-agent orchestration.
        """
        today = datetime.now().strftime("%Y-%m-%d")

        # ── Step 1: Quant metrics (direct, no LLM) ───────────────────────────
        quant = get_quant_metrics(ticker)
        logger.info("Quant metrics for %s: %s", ticker, quant)

        # ── Step 2: Fundamental data (direct SEC + yfinance) ─────────────────
        fundamental = self._fetch_fundamental(ticker)
        logger.info("Fundamental for %s: %s", ticker, fundamental.get("summary", "")[:100])

        # ── Step 3: News sentiment (direct Tavily API) ────────────────────────
        news = fetch_news_direct(ticker, company_name)
        logger.info("News for %s: signal=%s", ticker, news.get("signal"))

        # ── Step 4: Consensus ─────────────────────────────────────────────────
        consensus = self.calculate_consensus(
            quant_sig=quant.get("signal", "negative"),
            analyst_sig=fundamental.get("signal", "negative"),
            pulse_sig=news.get("signal", "negative"),
        )

        # ── Step 5: LLM synthesis only (not orchestration) ───────────────────
        synthesis_prompt = (
            f"You are a senior equity analyst. Based on the data below for {company_name} ({ticker}), "
            f"write a structured analysis. Return ONLY a JSON object starting with {{ and ending with }}.\n\n"
            f"QUANT DATA: {json.dumps(quant)}\n"
            f"FUNDAMENTAL DATA: {fundamental.get('summary', 'N/A')}\n"
            f"NEWS DATA: {news.get('summary', 'N/A')}\n"
            f"CONSENSUS: {json.dumps(consensus)}\n\n"
            f"Return JSON with these exact keys:\n"
            f"ticker, company_name, recommendation, confidence, "
            f"quant_signal, fundamental_signal, news_signal, "
            f"ann_return, ann_volatility, sharpe_ratio, current_price, market_cap, pe_ratio, "
            f"52w_high, 52w_low, "
            f"quant_reasoning, fundamental_reasoning, news_reasoning, overall_reasoning, "
            f"sec_url, news_urls, analysis_date"
        )

        try:
            llm_response = await Settings.llm.acomplete(synthesis_prompt)
            raw = str(llm_response).strip()
            logger.info("LLM synthesis for %s: %s", ticker, raw[:200])
            result = self._extract_json(raw)
        except Exception as exc:
            logger.error("LLM synthesis failed for %s: %s", ticker, exc)
            result = None

        if result:
            # Override with reliable direct data
            result["ticker"] = ticker
            result["company_name"] = company_name
            result["analysis_date"] = today
            result["recommendation"] = consensus.get("recommendation", result.get("recommendation", "HOLD"))
            result["confidence"] = consensus.get("confidence", result.get("confidence", "66%"))
            result["quant_signal"] = quant.get("signal", "N/A")
            result["fundamental_signal"] = fundamental.get("signal", "N/A")
            result["news_signal"] = news.get("signal", "N/A")
            result["ann_return"] = quant.get("ann_return", "N/A")
            result["ann_volatility"] = quant.get("ann_volatility", "N/A")
            result["sharpe_ratio"] = quant.get("sharpe_ratio", "N/A")
            result["current_price"] = quant.get("current_price", "N/A")
            result["market_cap"] = quant.get("market_cap", "N/A")
            result["pe_ratio"] = quant.get("pe_ratio", "N/A")
            result["52w_high"] = quant.get("52w_high", "N/A")
            result["52w_low"] = quant.get("52w_low", "N/A")
            result["sec_url"] = fundamental.get("sec_url", "")
            result["news_urls"] = news.get("urls", [])
            return result

        # If LLM synthesis fails, build from raw data
        return self._build_direct_result(
            ticker, company_name, quant, fundamental, news, consensus, today
        )

    def _fetch_fundamental(self, ticker: str) -> dict:
        """Fetch fundamental data directly — same logic as search_web_10k but callable from Python."""
        import requests
        sec_url = (
            f"https://www.sec.gov/cgi-bin/browse-edgar?"
            f"action=getcompany&CIK={ticker}&type=10-K&dateb=&owner=include&count=5"
        )
        headers = {"User-Agent": "PortfolioAI research@portfolioai.com"}
        sec_summary = ""

        try:
            tickers_resp = requests.get(
                "https://www.sec.gov/files/company_tickers.json",
                headers=headers, timeout=10,
            )
            if tickers_resp.status_code == 200:
                for entry in tickers_resp.json().values():
                    if entry.get("ticker", "").upper() == ticker.upper():
                        cik = str(entry["cik_str"]).zfill(10)
                        facts_resp = requests.get(
                            f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json",
                            headers=headers, timeout=15,
                        )
                        if facts_resp.status_code == 200:
                            facts = facts_resp.json()
                            us_gaap = facts.get("facts", {}).get("us-gaap", {})

                            def latest_val(concept):
                                usd = us_gaap.get(concept, {}).get("units", {}).get("USD", [])
                                annual = [x for x in usd if x.get("form") == "10-K"]
                                if annual:
                                    val = annual[-1].get("val", 0)
                                    return f"${val/1e9:.2f}B" if val >= 1e9 else f"${val/1e6:.2f}M"
                                return "N/A"

                            revenue = latest_val("Revenues") or latest_val("RevenueFromContractWithCustomerExcludingAssessedTax")
                            net_income = latest_val("NetIncomeLoss")
                            total_assets = latest_val("Assets")
                            sec_url = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}&type=10-K&dateb=&owner=include&count=5"
                            sec_summary = (
                                f"SEC filing data: Revenue={revenue}, "
                                f"Net Income={net_income}, Total Assets={total_assets}."
                            )
                        break
        except Exception as exc:
            logger.warning("SEC lookup failed for %s: %s", ticker, exc)

        try:
            info = yf.Ticker(ticker).info
            business = info.get("longBusinessSummary", "")[:500]
            sector = info.get("sector", "N/A")
            industry = info.get("industry", "N/A")
            pe = info.get("trailingPE", "N/A")
            profit_margin = info.get("profitMargins", "N/A")
            if isinstance(profit_margin, float):
                profit_margin = f"{profit_margin:.1%}"
            debt_equity = info.get("debtToEquity", "N/A")

            yf_summary = (
                f"Business: {business} "
                f"Sector: {sector}. Industry: {industry}. "
                f"P/E: {pe}. Profit Margin: {profit_margin}. Debt/Equity: {debt_equity}."
            )
            full_summary = (sec_summary + " " + yf_summary).strip()

            # Determine signal from financial health
            signal = "positive"
            if isinstance(pe, (int, float)) and pe > 50:
                signal = "negative"
            if isinstance(debt_equity, (int, float)) and debt_equity > 200:
                signal = "negative"
            if isinstance(info.get("profitMargins"), float) and info["profitMargins"] < 0:
                signal = "negative"

            return {"summary": full_summary, "signal": signal, "sec_url": sec_url}

        except Exception as exc:
            logger.error("yfinance fundamental failed for %s: %s", ticker, exc)
            summary = sec_summary or f"Could not retrieve fundamental data for {ticker}."
            return {"summary": summary, "signal": "negative", "sec_url": sec_url}

    def _build_direct_result(
        self, ticker, company_name, quant, fundamental, news, consensus, today
    ) -> dict:
        """Build a complete result directly from raw data when LLM synthesis fails."""
        rec = consensus.get("recommendation", "HOLD")
        conf = consensus.get("confidence", "66%")
        return {
            "ticker": ticker,
            "company_name": company_name,
            "recommendation": rec,
            "confidence": conf,
            "quant_signal": quant.get("signal", "N/A"),
            "fundamental_signal": fundamental.get("signal", "N/A"),
            "news_signal": news.get("signal", "N/A"),
            "ann_return": quant.get("ann_return", "N/A"),
            "ann_volatility": quant.get("ann_volatility", "N/A"),
            "sharpe_ratio": quant.get("sharpe_ratio", "N/A"),
            "current_price": quant.get("current_price", "N/A"),
            "market_cap": quant.get("market_cap", "N/A"),
            "pe_ratio": quant.get("pe_ratio", "N/A"),
            "52w_high": quant.get("52w_high", "N/A"),
            "52w_low": quant.get("52w_low", "N/A"),
            "quant_reasoning": (
                f"Sharpe ratio of {quant.get('sharpe_ratio','N/A')} with annual return of "
                f"{quant.get('ann_return','N/A')} and volatility of {quant.get('ann_volatility','N/A')}."
            ),
            "fundamental_reasoning": fundamental.get("summary", "N/A")[:300],
            "news_reasoning": news.get("summary", "N/A")[:300],
            "overall_reasoning": (
                f"Consensus recommendation of {rec} ({conf} confidence) based on: "
                f"Quant signal {quant.get('signal','N/A')}, "
                f"Fundamental signal {fundamental.get('signal','N/A')}, "
                f"News signal {news.get('signal','N/A')}."
            ),
            "sec_url": fundamental.get("sec_url", ""),
            "news_urls": news.get("urls", []),
            "analysis_date": today,
        }

    def _build_fallback_result(
        self, ticker: str, company_name: str, quant: dict, today: str, reasoning: str
    ) -> dict:
        """
        Build a best-effort result using direct quant data when the agent
        pipeline fails to return parseable JSON. Still shows useful data
        rather than an error screen.
        """
        quant_sig = quant.get("signal", "negative")
        sharpe = quant.get("sharpe_ratio", 0)
        try:
            sharpe_val = float(sharpe)
        except (ValueError, TypeError):
            sharpe_val = 0

        # Simple quant-only recommendation as fallback
        if sharpe_val > 1.5:
            rec, conf = "BUY", "66%"
        elif sharpe_val > 0.5:
            rec, conf = "HOLD", "66%"
        else:
            rec, conf = "SELL", "66%"

        return {
            "ticker": ticker,
            "company_name": company_name,
            "recommendation": rec,
            "confidence": conf,
            "quant_signal": quant_sig,
            "fundamental_signal": "unavailable",
            "news_signal": "unavailable",
            "ann_return": quant.get("ann_return", "N/A"),
            "ann_volatility": quant.get("ann_volatility", "N/A"),
            "sharpe_ratio": quant.get("sharpe_ratio", "N/A"),
            "current_price": quant.get("current_price", "N/A"),
            "market_cap": quant.get("market_cap", "N/A"),
            "pe_ratio": quant.get("pe_ratio", "N/A"),
            "52w_high": quant.get("52w_high", "N/A"),
            "52w_low": quant.get("52w_low", "N/A"),
            "quant_reasoning": (
                f"Sharpe ratio of {quant.get('sharpe_ratio','N/A')} with annual return of "
                f"{quant.get('ann_return','N/A')} and volatility of {quant.get('ann_volatility','N/A')}."
            ),
            "fundamental_reasoning": "Fundamental analysis unavailable — SEC data could not be retrieved.",
            "news_reasoning": "News sentiment unavailable — agent pipeline did not complete.",
            "overall_reasoning": (
                f"Recommendation based on quantitative metrics only (fundamental and news agents "
                f"did not return structured data). Sharpe ratio of {quant.get('sharpe_ratio','N/A')} "
                f"suggests a {rec} signal. Full analysis: {reasoning[:300]}"
            ),
            "sec_url": f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={ticker}&type=10-K",
            "news_urls": [],
            "analysis_date": today,
        }

    def generate_excel(self, results: list[dict]) -> str:
        """Write a richly formatted Excel report and return the file path."""
        path = os.path.join(self.output_dir, f"PortfolioAI_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        # ── Build flat rows ──────────────────────────────────────────────────
        rows = []
        for r in results:
            news_urls = r.get("news_urls", [])
            if isinstance(news_urls, str):
                try:
                    news_urls = json.loads(news_urls)
                except Exception:
                    news_urls = [news_urls]
            rows.append({
                "Ticker":               r.get("ticker", ""),
                "Company Name":         r.get("company_name", ""),
                "Analysis Date":        r.get("analysis_date", ""),
                "Recommendation":       r.get("recommendation", ""),
                "Confidence":           r.get("confidence", ""),
                "Quant Signal":         r.get("quant_signal", ""),
                "Fundamental Signal":   r.get("fundamental_signal", ""),
                "News Signal":          r.get("news_signal", ""),
                "Current Price":        r.get("current_price", ""),
                "Market Cap":           r.get("market_cap", ""),
                "P/E Ratio":            r.get("pe_ratio", ""),
                "52W High":             r.get("52w_high", ""),
                "52W Low":              r.get("52w_low", ""),
                "Annual Return":        r.get("ann_return", ""),
                "Annual Volatility":    r.get("ann_volatility", ""),
                "Sharpe Ratio":         r.get("sharpe_ratio", ""),
                "Overall Reasoning":    r.get("overall_reasoning", ""),
                "Quant Reasoning":      r.get("quant_reasoning", ""),
                "Fundamental Reasoning":r.get("fundamental_reasoning", ""),
                "News Reasoning":       r.get("news_reasoning", ""),
                "SEC Filing URL":       r.get("sec_url", ""),
                "News Source 1":        news_urls[0] if len(news_urls) > 0 else "",
                "News Source 2":        news_urls[1] if len(news_urls) > 1 else "",
                "News Source 3":        news_urls[2] if len(news_urls) > 2 else "",
            })

        df = pd.DataFrame(rows)
        df.to_excel(path, index=False, sheet_name="Analysis", engine="openpyxl")

        # ── Apply formatting ─────────────────────────────────────────────────
        wb = load_workbook(path)
        ws = wb["Analysis"]

        # Header style
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[1].height = 30

        # Colour-code recommendation column (col D = index 4)
        rec_col = 4
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            rec = ws.cell(row=row_idx, column=rec_col).value or ""
            if rec == "BUY":
                fill = PatternFill("solid", fgColor="C6EFCE")
                font_color = "276221"
            elif rec == "SELL":
                fill = PatternFill("solid", fgColor="FFC7CE")
                font_color = "9C0006"
            elif rec == "HOLD":
                fill = PatternFill("solid", fgColor="FFEB9C")
                font_color = "9C5700"
            else:
                fill = PatternFill("solid", fgColor="F2F2F2")
                font_color = "000000"

            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_idx, column=rec_col).fill = fill
            ws.cell(row=row_idx, column=rec_col).font = Font(
                bold=True, color=font_color
            )
            # Alternate row shading
            if row_idx % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = PatternFill("solid", fgColor="F7F9FC")

        # Make URLs clickable (FIXED: Added proper null/empty checks)
        url_cols = [21, 22, 23, 24]  # SEC URL + 3 news sources
        for col_idx in url_cols:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and cell.value != "":
                    try:
                        cell_str = str(cell.value).strip()
                        if cell_str and isinstance(cell_str, str) and cell_str.startswith("http"):
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0563C1", underline="single")
                    except (TypeError, AttributeError):
                        # Skip if conversion or string operation fails
                        pass

        # Column widths
        col_widths = {
            1: 10,  # Ticker
            2: 25,  # Company
            3: 14,  # Date
            4: 16,  # Recommendation
            5: 12,  # Confidence
            6: 14,  # Quant Signal
            7: 18,  # Fundamental Signal
            8: 14,  # News Signal
            9: 14,  # Price
            10: 14, # Mkt Cap
            11: 10, # PE
            12: 12, # 52W High
            13: 12, # 52W Low
            14: 14, # Ann Return
            15: 16, # Ann Vol
            16: 14, # Sharpe
            17: 50, # Overall Reasoning
            18: 40, # Quant Reasoning
            19: 40, # Fundamental Reasoning
            20: 40, # News Reasoning
            21: 35, # SEC URL
            22: 35, # News 1
            23: 35, # News 2
            24: 35, # News 3
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add a summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "PortfolioAI — Analysis Summary"
        ws_summary["A1"].font = Font(bold=True, size=14, color="1F3864")
        ws_summary["A3"] = "Generated:"
        ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_summary["A4"] = "Total Equities:"
        ws_summary["B4"] = len(results)
        buys = sum(1 for r in results if r.get("recommendation") == "BUY")
        sells = sum(1 for r in results if r.get("recommendation") == "SELL")
        holds = sum(1 for r in results if r.get("recommendation") == "HOLD")
        ws_summary["A5"] = "BUY:"
        ws_summary["B5"] = buys
        ws_summary["B5"].font = Font(color="276221", bold=True)
        ws_summary["A6"] = "HOLD:"
        ws_summary["B6"] = holds
        ws_summary["B6"].font = Font(color="9C5700", bold=True)
        ws_summary["A7"] = "SELL:"
        ws_summary["B7"] = sells
        ws_summary["B7"].font = Font(color="9C0006", bold=True)

        row = 9
        ws_summary.cell(row=row, column=1, value="Ticker").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value="Company").font = Font(bold=True)
        ws_summary.cell(row=row, column=3, value="Recommendation").font = Font(bold=True)
        ws_summary.cell(row=row, column=4, value="Confidence").font = Font(bold=True)
        ws_summary.cell(row=row, column=5, value="Sharpe Ratio").font = Font(bold=True)

        for i, r in enumerate(results, start=row + 1):
            ws_summary.cell(row=i, column=1, value=r.get("ticker", ""))
            ws_summary.cell(row=i, column=2, value=r.get("company_name", ""))
            rec = r.get("recommendation", "")
            ws_summary.cell(row=i, column=3, value=rec)
            color = "276221" if rec == "BUY" else "9C0006" if rec == "SELL" else "9C5700"
            ws_summary.cell(row=i, column=3).font = Font(bold=True, color=color)
            ws_summary.cell(row=i, column=4, value=r.get("confidence", ""))
            ws_summary.cell(row=i, column=5, value=r.get("sharpe_ratio", ""))

        for col in range(1, 6):
            ws_summary.column_dimensions[get_column_letter(col)].width = 20

        wb.save(path)
        logger.info("Excel report saved: %s", path)
        return path


# ---------------------------------------------------------------------------
# 6. DJIA DATA  (for left panel)
# ---------------------------------------------------------------------------
def get_djia_data() -> dict:
    """Fetch DJIA current performance data via yfinance."""
    try:
        dji = yf.Ticker("^DJI")
        hist = dji.history(period="5d")
        info = dji.info

        if hist.empty:
            return {"error": "Could not load DJIA data"}

        current = hist["Close"].iloc[-1]
        prev    = hist["Close"].iloc[-2]
        change  = current - prev
        pct     = (change / prev) * 100

        # 1-year performance
        hist_1y = dji.history(period="1y")
        ytd_return = ((hist_1y["Close"].iloc[-1] / hist_1y["Close"].iloc[0]) - 1) * 100

        # 30-day high/low
        hist_30 = dji.history(period="1mo")
        high_30 = hist_30["High"].max()
        low_30  = hist_30["Low"].min()

        # Build sparkline data (last 30 closes, normalised to % from start)
        closes = hist_1y["Close"].resample("W").last().tail(52).tolist()

        return {
            "current":    f"{current:,.2f}",
            "change":     f"{change:+,.2f}",
            "pct":        f"{pct:+.2f}%",
            "direction":  "up" if change >= 0 else "down",
            "ytd_return": f"{ytd_return:+.2f}%",
            "high_30":    f"{high_30:,.2f}",
            "low_30":     f"{low_30:,.2f}",
            "sparkline":  closes,
            "updated":    datetime.now().strftime("%H:%M ET"),
        }
    except Exception as exc:
        logger.error("DJIA fetch failed: %s", exc)
        return {"error": str(exc)}


def generate_djia_chart(years: int = 10) -> str | None:
    """
    Fetch DJIA history for the given number of years and generate
    an interactive Plotly HTML chart. Returns the file path or None on error.
    """
    try:
        dji = yf.Ticker("^DJI")
        period_map = {1: "1y", 2: "2y", 5: "5y", 10: "10y"}
        period = period_map.get(years, "10y")
        hist = dji.history(period=period)

        if hist.empty:
            return None

        hist = hist.reset_index()
        # Ensure Date column is datetime
        hist["Date"] = pd.to_datetime(hist["Date"])
        # Remove timezone info for Plotly compatibility
        if hist["Date"].dt.tz is not None:
            hist["Date"] = hist["Date"].dt.tz_localize(None)

        close   = hist["Close"]
        dates   = hist["Date"]
        color   = "#22c55e" if close.iloc[-1] >= close.iloc[0] else "#ef4444"

        # ── Build figure ────────────────────────────────────────────────────
        fig = go.Figure()

        # Filled area line
        fig.add_trace(go.Scatter(
            x=dates,
            y=close,
            mode="lines",
            name="DJIA",
            line=dict(color=color, width=2),
            fill="tozeroy",
            fillcolor=color.replace(")", ", 0.08)").replace("rgb", "rgba") if "rgb" in color
                      else f"rgba(34,197,94,0.08)" if color == "#22c55e"
                      else f"rgba(239,68,68,0.08)",
            hovertemplate="<b>%{x|%b %d, %Y}</b><br>DJIA: %{y:,.0f}<extra></extra>",
        ))

        # Add 200-day moving average
        if len(close) >= 200:
            ma200 = close.rolling(200).mean()
            fig.add_trace(go.Scatter(
                x=dates,
                y=ma200,
                mode="lines",
                name="200-Day MA",
                line=dict(color="#f59e0b", width=1.5, dash="dot"),
                hovertemplate="200D MA: %{y:,.0f}<extra></extra>",
            ))

        # Add 50-day moving average
        if len(close) >= 50:
            ma50 = close.rolling(50).mean()
            fig.add_trace(go.Scatter(
                x=dates,
                y=ma50,
                mode="lines",
                name="50-Day MA",
                line=dict(color="#a78bfa", width=1.5, dash="dash"),
                hovertemplate="50D MA: %{y:,.0f}<extra></extra>",
            ))

        # Annotations — start and end values
        pct_change = ((close.iloc[-1] - close.iloc[0]) / close.iloc[0]) * 100
        sign = "+" if pct_change >= 0 else ""

        fig.update_layout(
            title=dict(
                text=f"Dow Jones Industrial Average — {years}-Year Performance"
                     f"<br><sup>{sign}{pct_change:.1f}% over period  |  "
                     f"Current: {close.iloc[-1]:,.0f}  |  "
                     f"Updated: {datetime.now().strftime('%b %d, %Y %H:%M ET')}</sup>",
                font=dict(size=16, color="#f8fafc"),
                x=0,
            ),
            paper_bgcolor="#0f172a",
            plot_bgcolor="#0f172a",
            font=dict(color="#94a3b8", family="sans-serif"),
            xaxis=dict(
                showgrid=True,
                gridcolor="#1e293b",
                gridwidth=1,
                tickfont=dict(color="#94a3b8"),
                rangeslider=dict(visible=True, bgcolor="#1e293b", thickness=0.05),
                rangeselector=dict(
                    buttons=[
                        dict(count=6,  label="6M",  step="month", stepmode="backward"),
                        dict(count=1,  label="1Y",  step="year",  stepmode="backward"),
                        dict(count=3,  label="3Y",  step="year",  stepmode="backward"),
                        dict(count=5,  label="5Y",  step="year",  stepmode="backward"),
                        dict(step="all", label="All"),
                    ],
                    bgcolor="#1e293b",
                    activecolor="#334155",
                    font=dict(color="#94a3b8"),
                ),
            ),
            yaxis=dict(
                showgrid=True,
                gridcolor="#1e293b",
                gridwidth=1,
                tickfont=dict(color="#94a3b8"),
                tickformat=",.0f",
                side="right",
            ),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1,
                font=dict(color="#94a3b8"),
                bgcolor="rgba(0,0,0,0)",
            ),
            margin=dict(l=20, r=20, t=80, b=40),
            height=480,
            hovermode="x unified",
        )

        # Save as self-contained HTML file
        os.makedirs("./outputs", exist_ok=True)
        chart_path = "./outputs/djia_chart.html"
        fig.write_html(
            chart_path,
            include_plotlyjs="cdn",   # loads Plotly from CDN — keeps file small
            full_html=True,
            config={"displayModeBar": True, "scrollZoom": True},
        )
        return chart_path

    except Exception as exc:
        logger.error("DJIA chart generation failed: %s", exc)
        return None


def _djia_markdown(djia: dict) -> str:
    """Render DJIA data as clean Markdown — works reliably in all Chainlit versions."""
    if "error" in djia:
        return f"⚠️ Could not load DJIA data: {djia['error']}"

    arrow = "▲" if djia["direction"] == "up" else "▼"
    trend = "🟢" if djia["direction"] == "up" else "🔴"

    # ASCII sparkline from weekly closes (last 20 points)
    vals = djia.get("sparkline", [])
    spark = ""
    if vals and len(vals) > 1:
        mn, mx = min(vals), max(vals)
        rng = mx - mn or 1
        bars = ["▁", "▂", "▃", "▄", "▅", "▆", "▇", "█"]
        sample = vals[-20:]
        spark = "".join(bars[min(int((v - mn) / rng * 7), 7)] for v in sample)

#    return (
#        f"## 📊 Dow Jones Industrial Average\n\n"
#        f"### {trend} {djia['current']}  {arrow} {djia['change']} ({djia['pct']})\n\n"
#        f"`{spark}`\n\n"
#        f"| Metric | Value |\n"
#        f"|:---|---:|\n"
#        f"| 📈 1-Year Return | **{djia['ytd_return']}** |\n"
#        f"| 🔺 30D High | {djia['high_30']} |\n"
#        f"| 🔻 30D Low | {djia['low_30']} |\n"
#        f"| 🕐 Updated | {djia['updated']} |\n"
#    )


# ---------------------------------------------------------------------------
# 7. CHAINLIT UI
# ---------------------------------------------------------------------------
_bot: PortfolioBot | None = None


def get_bot() -> PortfolioBot:
    global _bot
    if _bot is None:
        _bot = PortfolioBot()
    return _bot


def _recommendation_emoji(rec: str) -> str:
    return {"BUY": "✅", "SELL": "🔴", "HOLD": "🟡"}.get(rec, "❓")


def _signal_emoji(sig: str) -> str:
    sig = str(sig).lower()
    if sig == "positive":  return "🟢"
    if sig == "negative":  return "🔴"
    return "⚪"


def _format_inline_result(r: dict) -> str:
    """
    Render a full equity result as rich inline Markdown.
    Everything displayed in the chat panel — no cards, no separate files shown first.
    """
    rec    = r.get("recommendation", "N/A")
    emoji  = _recommendation_emoji(rec)
    sig_q  = _signal_emoji(r.get("quant_signal", ""))
    sig_f  = _signal_emoji(r.get("fundamental_signal", ""))
    sig_n  = _signal_emoji(r.get("news_signal", ""))

    # Recommendation badge colour via unicode block
    badge  = {"BUY": "🟩 **BUY**", "SELL": "🟥 **SELL**", "HOLD": "🟨 **HOLD**"}.get(rec, f"**{rec}**")

    news_urls = r.get("news_urls", [])
    if isinstance(news_urls, str):
        try:    news_urls = json.loads(news_urls)
        except: news_urls = [news_urls]
    news_links = "\n".join(
        f"  - [📰 News Source {i+1}]({u})" for i, u in enumerate(news_urls) if u
    )
    sec_url  = r.get("sec_url", "")
    sec_link = f"  - [🏛️ SEC Filing]({sec_url})" if sec_url else ""
    sources  = "\n".join(filter(None, [sec_link, news_links])) or "  - No sources available"

    return f"""
---
# {emoji} {r.get('company_name', r.get('ticker', ''))} &nbsp; `{r.get('ticker', '')}`

> **{badge}** &nbsp;&nbsp; Confidence: **{r.get('confidence', 'N/A')}** &nbsp;&nbsp; Analysis Date: {r.get('analysis_date', '')}

---

## 📊 Agent Signals

| Agent | Signal | Reasoning |
|:---|:---:|:---|
| 📈 Quantitative | {sig_q} {r.get('quant_signal','N/A').title()} | {r.get('quant_reasoning', 'N/A')} |
| 🏛️ Fundamental | {sig_f} {r.get('fundamental_signal','N/A').title()} | {r.get('fundamental_reasoning', 'N/A')} |
| 📰 News Sentiment | {sig_n} {r.get('news_signal','N/A').title()} | {r.get('news_reasoning', 'N/A')} |

---

## 📈 Key Metrics

| Metric | Value | Metric | Value |
|:---|---:|:---|---:|
| **Current Price** | {r.get('current_price','N/A')} | **Market Cap** | {r.get('market_cap','N/A')} |
| **P/E Ratio** | {r.get('pe_ratio','N/A')} | **Sharpe Ratio** | {r.get('sharpe_ratio','N/A')} |
| **Annual Return** | {r.get('ann_return','N/A')} | **Annual Volatility** | {r.get('ann_volatility','N/A')} |
| **52-Week High** | {r.get('52w_high','N/A')} | **52-Week Low** | {r.get('52w_low','N/A')} |

---

## 💡 Overall Assessment

{r.get('overall_reasoning', 'N/A')}

---

## 🔗 Research Sources

{sources}

"""


@cl.on_chat_start
async def on_chat_start():
    # ── Fetch DJIA metrics and generate chart in parallel ────────────────────
    djia = get_djia_data()
    chart_path = generate_djia_chart(years=10)

    # ── Send Markdown metrics summary ────────────────────────────────────────
    await cl.Message(
        content=_djia_markdown(djia),
        author="Market Overview",
    ).send()

    # ── Attach interactive Plotly chart as downloadable/viewable HTML ────────
    if chart_path and os.path.exists(chart_path):
        await cl.Message(
            content=(
                "📈 **DJIA 10-Year Interactive Chart**\n\n"
                "_Open the file below for a fully interactive chart — "
                "zoom, pan, hover for daily values, and switch timeframes "
                "using the 6M / 1Y / 3Y / 5Y / All buttons. "
                "Includes 50-day and 200-day moving averages._"
            ),
            elements=[
                cl.File(
                    name="DJIA_10Year_Chart.html",
                    path=chart_path,
                    display="inline",
                )
            ],
            author="Market Overview",
        ).send()
    else:
        await cl.Message(
            content="⚠️ Could not generate DJIA chart — market data may be unavailable.",
            author="Market Overview",
        ).send()

    # ── Welcome message ──────────────────────────────────────────────────────
    await cl.Message(
        content=(
            "# 👋 Welcome to PortfolioAI\n\n"
            "I analyse equities using **three AI agents** — Quantitative, Fundamental (SEC filings), "
            "and News Sentiment — to deliver a **BUY / HOLD / SELL** recommendation "
            "with full reasoning displayed directly on screen.\n\n"
            "---\n\n"
            "## How to use\n\n"
            "| Input type | Example |\n"
            "|:---|:---|\n"
            "| Single ticker | `AAPL` |\n"
            "| Company name | `Apple` |\n"
            "| Multiple equities | `AAPL, Microsoft, NVDA` |\n"
            "| File upload | CSV or Excel with a ticker/company column |\n\n"
            "---\n\n"
            "_Each equity runs through all three agents independently. "
            "Results appear on screen as each analysis completes, "
            "followed by a downloadable Excel report with full reasoning and source links._\n\n"
            "**Type a ticker or company name below to get started ↓**"
        ),
        author="PortfolioAI",
    ).send()


@cl.on_message
async def on_message(message: cl.Message):
    bot = get_bot()
    tickers_to_analyse: list[tuple[str, str]] = []

    # ── Handle file uploads ──────────────────────────────────────────────────
    if message.elements:
        for element in message.elements:
            if hasattr(element, "path") and element.path:
                filename = element.name or ""
                if filename.lower().endswith((".csv", ".xlsx", ".xls")):
                    with open(element.path, "rb") as f:
                        raw_tokens = parse_file_upload(f.read(), filename)
                    await cl.Message(
                        content=f"📂 Found **{len(raw_tokens)}** equities in `{filename}`. Resolving tickers..."
                    ).send()
                    for token in raw_tokens:
                        ticker, name = resolve_ticker(token)
                        tickers_to_analyse.append((ticker, name))

    # ── Handle text input ────────────────────────────────────────────────────
    if message.content.strip():
        tokens = parse_ticker_input(message.content)
        if tokens:
            await cl.Message(
                content=f"🔍 Resolving **{len(tokens)}** equit{'y' if len(tokens)==1 else 'ies'}..."
            ).send()
            for token in tokens:
                ticker, name = resolve_ticker(token)
                tickers_to_analyse.append((ticker, name))

    if not tickers_to_analyse:
        await cl.Message(
            content=(
                "⚠️ I couldn't find any equities in your input.\n\n"
                "Please enter ticker symbols or company names (comma-separated), "
                "or upload a CSV/Excel file with a ticker column."
            )
        ).send()
        return

    # Deduplicate
    seen, unique = set(), []
    for t, n in tickers_to_analyse:
        if t not in seen:
            seen.add(t)
            unique.append((t, n))
    tickers_to_analyse = unique

    await cl.Message(
        content=(
            f"🚀 Starting analysis of **{len(tickers_to_analyse)}** "
            f"equit{'y' if len(tickers_to_analyse)==1 else 'ies'}: "
            f"{', '.join(f'`{t}`' for t, _ in tickers_to_analyse)}\n\n"
            "_Running Quant, Fundamental, and News agents for each equity — "
            "results will appear below as they complete..._"
        )
    ).send()

    # ── Analyse each equity and display inline ───────────────────────────────
    all_results = []
    for ticker, company_name in tickers_to_analyse:
        async with cl.Step(name=f"⏳ Analysing {ticker} — {company_name}") as step:
            step.output = "Running Quant → Fundamental → News → Consensus..."
            result = await bot.analyse_ticker(ticker, company_name)
            all_results.append(result)
            rec = result.get("recommendation", "ERROR")
            step.output = f"{'✅' if rec != 'ERROR' else '⚠️'} {ticker} complete — {rec}"

        # Display full inline result immediately
        await cl.Message(
            content=_format_inline_result(result),
            author=f"PortfolioAI — {company_name}",
        ).send()

    # ── Generate and attach Excel ────────────────────────────────────────────
    async with cl.Step(name="📊 Generating Excel report...") as step:
        excel_path = bot.generate_excel(all_results)
        step.output = "Report ready."

    buys  = sum(1 for r in all_results if r.get("recommendation") == "BUY")
    holds = sum(1 for r in all_results if r.get("recommendation") == "HOLD")
    sells = sum(1 for r in all_results if r.get("recommendation") == "SELL")

    elements = [
        cl.File(
            name=os.path.basename(excel_path),
            path=excel_path,
            display="inline",
        )
    ]

    await cl.Message(
        content=(
            f"---\n\n"
            f"## 📋 Analysis Complete\n\n"
            f"**{len(all_results)}** equit{'y' if len(all_results)==1 else 'ies'} analysed &nbsp;|&nbsp; "
            f"✅ **{buys} BUY** &nbsp;|&nbsp; "
            f"🟡 **{holds} HOLD** &nbsp;|&nbsp; "
            f"🔴 **{sells} SELL**\n\n"
            f"The Excel report below contains all results, full reasoning, "
            f"and clickable links to SEC filings and news sources."
        ),
        elements=elements,
        author="PortfolioAI",
    ).send()
